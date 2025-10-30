"""
Utilidades de Exportación para FacturIA 2.1.0
Exporta transacciones a Excel y PDF con formato profesional
"""
import pandas as pd
from datetime import datetime
from io import BytesIO
from typing import List, Dict
import sys


def exportar_a_excel(transacciones: List, nombre_archivo: str = None) -> BytesIO:
    """
    Exporta transacciones a Excel con formato profesional

    Args:
        transacciones: Lista de objetos Transaccion
        nombre_archivo: Nombre base del archivo (opcional)

    Returns:
        BytesIO con el archivo Excel
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        # Fallback: exportar como CSV sin formato
        return exportar_a_csv(transacciones)

    # Crear DataFrame
    data = []
    for t in transacciones:
        data.append({
            "ID": t.id,
            "Fecha": t.fecha_transaccion.strftime('%Y-%m-%d') if t.fecha_transaccion else "",
            "Tipo": t.tipo.value.upper() if t.tipo else "",
            "Categoría": t.categoria or "",
            "Monto": float(t.monto) if t.monto else 0,
            "Persona": t.persona or "General",
            "Emisor/Receptor": t.emisor_receptor or "",
            "Descripción": t.descripcion or "",
            "Número Comprobante": t.numero_comprobante or "",
            "Origen": t.origen.value.upper() if hasattr(t, 'origen') and t.origen else "",
            "Procesado por IA": "Sí" if t.procesado_por_ia else "No",
            "Requiere Revisión": "Sí" if t.requiere_revision else "No",
            "Editado Manualmente": "Sí" if t.editado_manualmente else "No"
        })

    df = pd.DataFrame(data)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    # Estilos
    header_fill = PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Agregar título
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = "FacturIA 2.1.0 - Reporte de Transacciones"
    title_cell.font = Font(bold=True, size=16, color="00CC66")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Agregar metadata
    ws['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Total de transacciones: {len(transacciones)}"

    # Agregar headers (fila 5)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Agregar datos
    for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=6):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border

            # Formato especial para montos
            if col_idx == 5:  # Columna Monto
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    # Ajustar ancho de columnas
    column_widths = {
        'A': 8, 'B': 12, 'C': 10, 'D': 20, 'E': 15,
        'F': 15, 'G': 25, 'H': 30, 'I': 20, 'J': 12,
        'K': 15, 'L': 18, 'M': 18
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Agregar filtros
    ws.auto_filter.ref = f"A5:M{len(transacciones) + 5}"

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def exportar_a_csv(transacciones: List) -> BytesIO:
    """
    Exporta transacciones a CSV simple

    Args:
        transacciones: Lista de objetos Transaccion

    Returns:
        BytesIO con el archivo CSV
    """
    data = []
    for t in transacciones:
        data.append({
            "ID": t.id,
            "Fecha": t.fecha_transaccion.strftime('%Y-%m-%d') if t.fecha_transaccion else "",
            "Tipo": t.tipo.value.upper() if t.tipo else "",
            "Categoría": t.categoria or "",
            "Monto": float(t.monto) if t.monto else 0,
            "Persona": t.persona or "General",
            "Emisor/Receptor": t.emisor_receptor or "",
            "Descripción": t.descripcion or "",
            "Número Comprobante": t.numero_comprobante or ""
        })

    df = pd.DataFrame(data)

    output = BytesIO()
    df.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)

    return output


def exportar_a_pdf(transacciones: List, stats: Dict = None) -> BytesIO:
    """
    Exporta transacciones a PDF con formato profesional

    Args:
        transacciones: Lista de objetos Transaccion
        stats: Diccionario con estadísticas (opcional)

    Returns:
        BytesIO con el archivo PDF
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    except ImportError:
        # Si reportlab no está disponible, devolver None
        return None

    output = BytesIO()

    # Crear documento
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    # Estilos
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#00CC66'),
        spaceAfter=12,
        alignment=TA_CENTER
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=12,
        alignment=TA_CENTER
    )

    # Elementos del documento
    elements = []

    # Título
    elements.append(Paragraph("FacturIA 2.1.0", title_style))
    elements.append(Paragraph("Reporte de Transacciones Financieras", subtitle_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
    elements.append(Spacer(1, 0.3*inch))

    # Estadísticas (si se proporcionan)
    if stats:
        stats_data = [
            ["Concepto", "Valor"],
            ["Total Transacciones", str(stats.get('total_transacciones', 0))],
            ["Total Ingresos", f"${stats.get('total_ingresos', 0):,.2f}"],
            ["Total Egresos", f"${stats.get('total_egresos', 0):,.2f}"],
            ["Balance", f"${stats.get('balance', 0):,.2f}"]
        ]

        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00CC66')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(stats_table)
        elements.append(Spacer(1, 0.3*inch))

    # Tabla de transacciones
    elements.append(Paragraph("Detalle de Transacciones", styles['Heading2']))
    elements.append(Spacer(1, 0.2*inch))

    # Preparar datos
    table_data = [["ID", "Fecha", "Tipo", "Categoría", "Monto", "Persona"]]

    for t in transacciones[:100]:  # Limitar a 100 para no hacer el PDF muy grande
        table_data.append([
            str(t.id),
            t.fecha_transaccion.strftime('%Y-%m-%d') if t.fecha_transaccion else "",
            t.tipo.value.upper() if t.tipo else "",
            t.categoria[:15] if t.categoria else "",
            f"${t.monto:,.2f}",
            t.persona[:12] if t.persona else "General"
        ])

    # Crear tabla
    trans_table = Table(table_data, colWidths=[0.5*inch, 1*inch, 0.8*inch, 1.5*inch, 1.2*inch, 1.2*inch])
    trans_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00CC66')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))

    elements.append(trans_table)

    # Nota al pie
    if len(transacciones) > 100:
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(
            f"Nota: Se muestran las primeras 100 de {len(transacciones)} transacciones",
            subtitle_style
        ))

    # Construir PDF
    doc.build(elements)

    output.seek(0)
    return output


if __name__ == "__main__":
    print("✅ Módulo de exportación cargado correctamente")
    print("Funciones disponibles:")
    print("  - exportar_a_excel(transacciones)")
    print("  - exportar_a_csv(transacciones)")
    print("  - exportar_a_pdf(transacciones, stats)")
